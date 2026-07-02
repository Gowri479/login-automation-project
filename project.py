import time
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.chrome.service import Service

EXCEL_FILE = "accounts.xlsx"

# Common attribute values seen across many login forms.
# The script tries each one in order until it finds a match.
USERNAME_GUESSES = [
    (By.ID, "LoginForm_username"),
    (By.ID, "username"), (By.NAME, "username"),
    (By.ID, "user-name"), (By.NAME, "user-name"),
    (By.ID, "email"), (By.NAME, "email"),
    (By.ID, "user"), (By.NAME, "user"),
    (By.CSS_SELECTOR, "input[type='email']"),
    (By.CSS_SELECTOR, "input[autocomplete='username']"),
]

PASSWORD_GUESSES = [
    (By.ID, "LoginForm_password"),
    (By.ID, "password"), (By.NAME, "password"),
    (By.ID, "pass"), (By.NAME, "pass"),
    (By.CSS_SELECTOR, "input[type='password']"),
]

BUTTON_GUESSES = [
    (By.XPATH, "//button[@type='submit']"),
    (By.ID, "submit"), (By.ID, "login-button"),
    (By.XPATH, "//button[contains(translate(text(),'LOGIN','login'),'login')]"),
    (By.XPATH, "//input[@type='submit']"),
]


def find_first(driver, guesses):
    for by, value in guesses:
        try:
            el = driver.find_element(by, value)
            if el.is_displayed():
                return el
        except NoSuchElementException:
            continue
    return None


def read_accounts(path):
    wb = load_workbook(path)
    sheet = wb.active
    accounts = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        url, username, password = row[0], row[1], row[2]
        if url and username and password:
            accounts.append({"url": url, "username": username, "password": password})
    return accounts


def login(account, driver, wait):
    driver.get(account["url"])
    time.sleep(1.5)  # let the page settle

    user_box = find_first(driver, USERNAME_GUESSES)
    pass_box = find_first(driver, PASSWORD_GUESSES)

    if not user_box or not pass_box:
        print(f"Could not find login fields: {account['username']} ({account['url']})")
        return False

    user_box.clear()
    user_box.send_keys(account["username"])
    pass_box.clear()
    pass_box.send_keys(account["password"])

    button = find_first(driver, BUTTON_GUESSES)
    if button:
        button.click()
    else:
        pass_box.submit()  # fallback: submit the form directly

    time.sleep(2)

    # Generic success check: URL changed away from a "login" page,
    # or the password field is no longer visible.
    still_on_login = "login" in driver.current_url.lower() or "signin" in driver.current_url.lower()
    password_still_visible = find_first(driver, PASSWORD_GUESSES) is not None

    if not still_on_login or not password_still_visible:
        print(f"Successfully logged in: {account['username']} ({account['url']})")
        return True
    else:
        print(f"Login failed or could not be verified: {account['username']} ({account['url']})")
        return False


def main():
    accounts = read_accounts(EXCEL_FILE)
    print(f"Accounts found: {len(accounts)}")
    if not accounts:
        print("No accounts found in the Excel file.")
        return

    service = Service(executable_path=r"C:\WebDrivers\chromedriver.exe")
    driver = webdriver.Chrome(service=service)
    wait = WebDriverWait(driver, 15)

    try:
        for account in accounts:
            driver.delete_all_cookies()
            login(account, driver, wait)
    finally:
        driver.quit()


if __name__ == "__main__":
    main()